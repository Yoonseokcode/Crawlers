from selenium import webdriver
from bs4 import BeautifulSoup
import time
import openpyxl
import random
from openpyxl.styles import alignment
import os
from datetime import datetime
import sys

chrome_options=webdriver.ChromeOptions()
chrome_options.add_argument('headless')
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}  # 크롤러 세팅
path=os.getcwd()+"/chromedriver.exe"        #실행파일과 같은 위치에 크롬드라이버가 있으면 작동 가능     #나중에 상황에 맞게 수정할 것
if os.path.isfile(path)==False:
    print('크롬 드라이버가 존재하지 않습니다. 본 폴더에 크롬 드라이버를 넣어 주세요.')
    os.system('pause')
driver=webdriver.Chrome(path,chrome_options=chrome_options)     #Chrome driver&header 세팅 집합

excel_file=openpyxl.load_workbook('StremCrawlingResult.xlsm', read_only=False,keep_vba=True)
excel_sheet1=excel_file['Catalog']
excel_sheet2=excel_file['Price']
excel_sheet1.protection.sheet=False
excel_sheet1.protection.sheet=False         #엑셀파일 불러오기

LastCAT_Num=excel_sheet2.cell(row=excel_sheet2.max_row,column=1).value
i=2
while excel_sheet1.cell(row=i,column=3).value!=LastCAT_Num:     #Catalog 시트에서 마지막으로 진행된 CAT_Num 위치 찾기
    i+=1

counter=i+1     #마지막으로 진행된 CAT_Num 바로 다음 것부터 재시작

if counter>excel_sheet1.max_row:
    print("더 이상 크롤링할 내용이 없습니다. 잠시 후 프로그램이 자동으로 종료됩니다.")
    time.sleep(3)
    sys.exit()

number=int(input('현재까지 진행된 번호: '+str(counter-2)+'\n추가로 진행할 개수 입력: '))

if number+i>excel_sheet1.max_row:
    print('입력 범위가 전체 데이터 수를 초과했으므로 도중에 프로그램이 종료됩니다.')
    number=excel_sheet1.max_row-i

EndNum=counter+number

while counter<=EndNum-1:      #마지막으로 진행된 번호의 다음 번호부터 입력된 숫자만큼 추가 진행
    randomtime = random.randrange(1, 4)
    CAT_Num = excel_sheet1.cell(row=counter, column=3).value
    time.sleep(randomtime)
    driver.get('https://www.strem.com/catalog/v/' + CAT_Num + '/')
    time.sleep(0.5)
    if counter == i+1:          #처음 실행 시 South Korea 선택하는 부분
        driver.find_element_by_xpath("//option[@value='South Korea']").click()
        driver.find_element_by_xpath('//*[@id="country_selection_form"]/div/div[1]/input[1]').click()
    data = driver.page_source
    soup = BeautifulSoup(data, 'html.parser')
    Info = soup.select('#purchase_form > table > tbody > tr')
    Line = []
    for tr in Info:
        Size = tr.select_one('td.size')
        if Size is not None:
            Size = Size.text

        Price = tr.select_one('td.price')
        if Price is not None:
            Price = Price.text
            PriceArr = Price.split('\n')  # \n값 제거 코드
            Price = PriceArr[1]
            PriceArr = Price.split(' ')  # 뒤에 남는 공간 제거 코드
            Price = PriceArr[0]
            PriceArr=Price.split('$')
            Price=PriceArr[1]
            if Price.find(',')!=-1:         #4자릿수 가격은 형변환이 안되므로 사이의 반점을 제거하는 코드를 삽입함
                PriceArr=Price.split(',')
                Price=PriceArr[0]+PriceArr[1]
            Price=float(Price)

        Availability = tr.select_one('td.availability > div.summary')
        if Availability is not None:
            Availability = Availability.text
        Time=datetime.now()
        Line = [CAT_Num, Size, Price, Availability,Time]
        excel_sheet2.append(Line)
    print(str(counter-1) + '번 페이지 크롤링 완료')
    if counter==excel_sheet1.max_row:
        break
    counter += 1

driver.quit()

print('후처리 작업 중...')
i=EndNum-number                                                                # 중간에 들어가는 None값 제거&날짜 형태 변환 코드
while i <= excel_sheet2.max_row:
    if excel_sheet2.cell(row=i, column=4).value in [None]:
        excel_sheet2.delete_rows(i)
    elif str(excel_sheet2.cell(row=i,column=4).value).split(' ')[0] == 'Available':
        stringcell=str(excel_sheet2.cell(row=i,column=4).value)
        cellArray=stringcell.split(' ')
        stringcell=cellArray[2]
        datecell=datetime.strptime(stringcell,'%d-%b-%Y')
        datecell=datetime.date(datecell)
        excel_sheet2.cell(row=i,column=4).value=datecell
        i+=1
    else:
        i += 1

excel_sheet2.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')    # A1셀 가운데 정렬
excel_sheet2.cell(row=1, column=2).alignment = openpyxl.styles.Alignment(horizontal='center')    # B1셀 가운데 정렬
excel_sheet2.cell(row=1, column=3).alignment = openpyxl.styles.Alignment(horizontal='center')    # C1셀 가운데 정렬
excel_sheet2.cell(row=1, column=4).alignment = openpyxl.styles.Alignment(horizontal='center')    # D1셀 가운데 정렬       #엑셀 1열 가운데정렬 집합

excel_sheet1.protection.sheet=True
excel_sheet2.protection.sheet=True
excel_file.save('StremCrawlingResult.xlsm')
excel_file.close()
