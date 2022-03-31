from selenium import webdriver
from bs4 import BeautifulSoup
import time
import openpyxl
import random
from openpyxl.styles import alignment
import os
from datetime import datetime

chrome_options=webdriver.ChromeOptions()
chrome_options.add_argument('headless')
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}  # 크롤러 세팅
path=os.getcwd()+"/chromedriver.exe"        #실행파일과 같은 위치에 크롬드라이버가 있으면 작동 가능     #나중에 상황에 맞게 수정할 것
if os.path.isfile(path)==False:
    print('크롬 드라이버가 존재하지 않습니다. 본 폴더에 크롬 드라이버를 넣어 주세요.')
    os.system('pause')
driver=webdriver.Chrome(path,chrome_options=chrome_options)     #Chrome driver&header 세팅 집합

excel_file=openpyxl.load_workbook('StremCrawlingResult.xlsm',read_only=False,keep_vba=True)
excel_sheet1=excel_file['Catalog']
excel_sheet2=excel_file.create_sheet('Price')
excel_sheet2.column_dimensions['A'].width = 15
excel_sheet2.column_dimensions['B'].width = 13
excel_sheet2.column_dimensions['C'].width = 10
excel_sheet2.column_dimensions['D'].width = 11
excel_sheet2.column_dimensions['E'].width = 18
excel_sheet2.append(['Catalog Number', 'Size', 'Price', 'Stock','Time'])       #Price sheet 세팅 집합

counter=2

while counter<=101:            #우선 100번째 데이터까지 크롤링을 진행
    randomtime = random.randrange(1, 4)
    CAT_Num=excel_sheet1.cell(row=counter,column=3).value
    time.sleep(randomtime)
    driver.get('https://www.strem.com/catalog/v/' + CAT_Num + '/')
    time.sleep(0.5)
    if counter==2:
        driver.find_element_by_xpath("//option[@value='South Korea']").click()      #south korea 선택
        driver.find_element_by_xpath('//*[@id="country_selection_form"]/div/div[1]/input[1]').click()       #select 버튼 클릭
    data=driver.page_source
    soup = BeautifulSoup(data, 'html.parser')
    Info = soup.select('#purchase_form > table > tbody > tr')
    Line = []
    for tr in Info:
        Size = tr.select_one('td.size')
        if Size is not None:
            Size = Size.text

        Price = tr.select_one('td.price')
        if Price is not None:
            Price = Price.text
            PriceArr = Price.split('\n')  # \n값 제거 코드
            Price = PriceArr[1]
            PriceArr = Price.split(' ')  # 뒤에 남는 공간 제거 코드
            Price = PriceArr[0]
            PriceArr=Price.split('$')
            Price=PriceArr[1]
            if Price.find(',')!=-1:         #4자릿수 가격은 형변환이 안되므로 사이의 반점을 제거하는 코드를 삽입함
                PriceArr=Price.split(',')
                Price=PriceArr[0]+PriceArr[1]
            Price=float(Price)

        Availability = tr.select_one('td.availability > div.summary')
        if Availability is not None:
            Availability = Availability.text
        Time = datetime.now()
        Line = [CAT_Num,Size, Price, Availability,Time]
        excel_sheet2.append(Line)
    print(str(counter-1)+'번 페이지 크롤링 완료')
    counter += 1
driver.quit()

print('후처리 작업 중...')

i = 1                                                                # 중간에 들어가는 None값과 재고없는 데이터의 앞부분을 제거하는 부분
while i <= excel_sheet2.max_row:
    if excel_sheet2.cell(row=i, column=4).value in [None]:
        excel_sheet2.delete_rows(i)
    elif str(excel_sheet2.cell(row=i, column=4).value).split(' ')[0] == 'Available':
        stringcell = str(excel_sheet2.cell(row=i, column=4).value)
        cellArray = stringcell.split(' ')
        stringcell = cellArray[2]
        datecell = datetime.strptime(stringcell, '%d-%b-%Y')
        datecell = datetime.date(datecell)
        excel_sheet2.cell(row=i, column=4).value = datecell
        i += 1
    else:
        i += 1

excel_sheet2.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')    # A1셀 가운데 정렬
excel_sheet2.cell(row=1, column=2).alignment = openpyxl.styles.Alignment(horizontal='center')    # B1셀 가운데 정렬
excel_sheet2.cell(row=1, column=3).alignment = openpyxl.styles.Alignment(horizontal='center')    # C1셀 가운데 정렬
excel_sheet2.cell(row=1, column=4).alignment = openpyxl.styles.Alignment(horizontal='center')    # D1셀 가운데 정렬
excel_sheet2.cell(row=1, column=5).alignment = openpyxl.styles.Alignment(horizontal='center')    #엑셀 1열 가운데정렬 집합

excel_sheet1.protection.sheet=True
excel_sheet2.protection.sheet=True
excel_file.save('StremCrawlingResult.xlsm')
excel_file.close()
