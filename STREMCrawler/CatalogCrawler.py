import requests
import random
import time
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import alignment

excel_file = openpyxl.Workbook()
excel_sheet1 = excel_file.active
excel_sheet1.title='Catalog'
excel_sheet1.column_dimensions['A'].width = 38
excel_sheet1.column_dimensions['B'].width = 36
excel_sheet1.column_dimensions['C'].width = 15
excel_sheet1.column_dimensions['D'].width = 140
excel_sheet1.column_dimensions['E'].width = 15
excel_sheet1.append(['대분류', '중분류', 'Catalog Number', 'Description', 'CAS Number'])               #엑셀 파일 세팅 집합

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}  # 크롤러 세팅

counter = 1

while counter <= 109:
    randomTime = random.randrange(1, 4)
    counterstr = str(counter)

    time.sleep(randomTime)

    data = requests.get('https://www.strem.com/catalog/gl/' + counterstr + '/', headers=headers)

    soup = BeautifulSoup(data.text, 'html.parser')

    Material = soup.select('#primary_content > div.product_section > div:nth-child(1) > span.category > a')
    Material = str(Material)
    if Material != '[]':
        MaterialArr = Material.split('>')
        Material = MaterialArr[1]
        MaterialArr = Material.split('<')
        Material = MaterialArr[0]
        if counter == 93:
            Material = Material.replace('&amp;', '&')           # 대분류 받아오기

    Compounds = soup.select('#catalog_section_0 > table.product_list.list > tbody > tr')
    Line = []

    for tr in Compounds:
        Catalog_Num = tr.select_one('td.catalog_number > a')
        if Catalog_Num is not None:
            Catalog_Num = Catalog_Num.text

        Description = tr.select_one('td.description > a')
        if Description is not None:
            Description = Description.text

        CAS_Num = tr.select_one('td.cas > a')
        if CAS_Num is not None:
            CAS_Num = CAS_Num.text

        Line = [Material, 'Compounds', Catalog_Num, Description, CAS_Num]
        excel_sheet1.append(Line)

    ElemantalForms = soup.select('#catalog_section_1 > table > tbody > tr')

    for tr in ElemantalForms:
        Catalog_Num = tr.select_one('td.catalog_number > a')
        if Catalog_Num is not None:
            Catalog_Num = Catalog_Num.text

        Description = tr.select_one('td.description > a')
        if Description is not None:
            Description = Description.text

        CAS_Num = tr.select_one('td.cas > a')
        if CAS_Num is not None:
            CAS_Num = CAS_Num.text
        Line = [Material, 'Elemental forms', Catalog_Num, Description, CAS_Num]
        excel_sheet1.append(Line)
    print(counterstr + '번 페이지 크롤링 완료')
    counter += 1

print('후처리 작업 중...')
i=2
while i <= excel_sheet1.max_row:
    if excel_sheet1.cell(row=i, column=4).value in [None] or excel_sheet1.cell(row=i, column=1).value in [None]:        # 중간에 들어가는 None값&2번 입력되는 Tin값들 제거 코드
        excel_sheet1.delete_rows(i)
    else:
        i+=1

excel_sheet1.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')    # A1셀 가운데 정렬
excel_sheet1.cell(row=1, column=2).alignment = openpyxl.styles.Alignment(horizontal='center')    # B1셀 가운데 정렬
excel_sheet1.cell(row=1, column=3).alignment = openpyxl.styles.Alignment(horizontal='center')    # C1셀 가운데 정렬
excel_sheet1.cell(row=1, column=4).alignment = openpyxl.styles.Alignment(horizontal='center')    # D1셀 가운데 정렬
excel_sheet1.cell(row=1, column=5).alignment = openpyxl.styles.Alignment(horizontal='center')    # E1셀 가운데 정렬        #엑셀 1열 가운데정렬 코드 집합

excel_file.save('StremCrawlingResult.xlsm')
excel_file1=openpyxl.load_workbook('StremCrawlingResult.xlsm',read_only=False,keep_vba=True)
excel_file1.save('StremCrawlingResult.xlsm')
excel_file.close()
excel_file1.close()

import StremPandasProcess
StremPandasProcess

print('Catalog 시트 작성 완료. Price 시트 작성을 시작합니다...')