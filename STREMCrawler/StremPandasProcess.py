import pandas as pd
import openpyxl
from openpyxl.styles import alignment

df=pd.read_excel('StremCrawlingResult.xlsm')

df=df.drop_duplicates(subset='Catalog Number')
df=df.sort_values(by=['대분류','중분류'],ascending=True)

df.to_excel('StremCrawlingResult.xlsm',index=False,header=False)

excel_file=openpyxl.load_workbook('StremCrawlingResult.xlsm',read_only=False,keep_vba=True)
excel_sheet1 = excel_file.active
excel_sheet1.title='Catalog'
excel_sheet1.column_dimensions['A'].width = 38
excel_sheet1.column_dimensions['B'].width = 36
excel_sheet1.column_dimensions['C'].width = 15
excel_sheet1.column_dimensions['D'].width = 140
excel_sheet1.column_dimensions['E'].width = 15
excel_sheet1.insert_rows(1)
excel_sheet1.cell(row=1,column=1).value='대분류'
excel_sheet1.cell(row=1,column=2).value='중분류'
excel_sheet1.cell(row=1,column=3).value='Catalog Number'
excel_sheet1.cell(row=1,column=4).value='Description'
excel_sheet1.cell(row=1,column=5).value='CAS Number'
i=1
while i<=5:
    excel_sheet1.cell(row=1, column=i).alignment = openpyxl.styles.Alignment(horizontal='center')
    i+=1                    #엑셀 시트 변환 코드 집합
excel_file.save('StremCrawlingResult.xlsm')
excel_file.close()