from selenium import webdriver
from bs4 import BeautifulSoup
import time
import openpyxl
import os
import math
import random
from fake_useragent import UserAgent
from selenium.webdriver.firefox.options import Options
import subprocess
import EnaminePandasProcess

#tor브라우저가 있어야 실행 가능

path=os.getcwd()+"/geckodriver.exe"     #드라이버 경로
if not os.path.isfile(path):
    print('Gecko 드라이버가 존재하지 않습니다. 본 폴더에 Gecko 드라이버를 넣어 주세요.')
    os.system('pause')

if not os.path.isfile(os.getcwd() + r'\Tor Browser\Browser\TorBrowser\Tor\tor.exe'):       #상황에 맞게 수정할 것
    print('Tor Browser가 존재하지 않습니다. 본 폴더에 Tor Browser 폴더를 넣어 주세요.')
    os.system('pause')        # 파일 경로 및 파일 유무 체크

options=Options()
options.add_argument("--headless")
profile = webdriver.FirefoxProfile()

torexe = subprocess.Popen(os.getcwd()+r'\Tor Browser\Browser\TorBrowser\Tor\tor.exe')      #상황에 맞게 수정할 것
profile.set_preference("network.proxy.type", 1)
profile.set_preference("network.proxy.socks", "127.0.0.1")
profile.set_preference("network.proxy.socks_port", 9050)        # ip 변경 부분


ua=UserAgent()
profile.set_preference('general.useragent.override',str(ua.random))        # UserAgent 랜덤변경 부분

profile.update_preferences()

driver = webdriver.Firefox(options=options, executable_path=path,firefox_profile=profile, service_log_path=os.devnull)

filepath=os.getcwd()+'/EnamineCrawlingResult.xlsm'

if os.path.isfile(filepath):
    excel_file = openpyxl.load_workbook('EnamineCrawlingResult.xlsm', read_only=False, keep_vba=True)
    excel_sheet1 = excel_file['Catalog']
    excel_sheet1.protection.sheet=False
else:
    excel_file = openpyxl.Workbook()
    excel_sheet1 = excel_file.active
    excel_sheet1.title='Catalog'
    excel_sheet1.append(['Name', 'CAS Number', 'Catalog Number', 'Purity', 'Price/Quantity','Stock'])         #엑셀 파일 세팅 집합

driver.get('https://www.enaminestore.com/search')

while True:
    InputCasNum = input('CAS 번호 입력: ')
    if '\n' in InputCasNum:
        InputCasNum = InputCasNum.replace('\n', '')
    driver.find_element_by_class_name('c-search__input').send_keys(InputCasNum)
    try:
        driver.find_element_by_class_name('c-search__button').click()
        time.sleep(random.randrange(7, 10))
        data = driver.page_source
        soup = BeautifulSoup(data, 'html.parser')

        ProductInfo='#commongrid > table > tbody > tr:nth-child(1) > td:nth-child(2) > '
        Availability='#commongrid > table > tbody > tr:nth-child(1) > td:nth-child(3) > '
        Line=[]
        PriceAndUnitList=[]
        i=0
        j=0
        PriceAndUnitStr=''

        CatalogID=soup.select_one(ProductInfo + 'div:nth-child(1) > span.catdata1.cdatamarker')     #Cat_No
        if CatalogID is not None:
            CatalogID=CatalogID.text

        Name=soup.select_one(ProductInfo + 'div:nth-child(3) > span.catdata1')      #Name
        if Name is not None:
            Name=Name.text

        CASNum=soup.select_one(ProductInfo + 'div:nth-child(5) > span.catdata1')        #Cas_No
        if CASNum is not None:
            CASNum=CASNum.text

        # Purity
        Purity=soup.select_one(Availability + 'span:nth-child(1)')
        if Purity is not None:
            Purity=Purity.text
            PurityArr=Purity.split(' ')
            Purity=PurityArr[1]
            if Purity.find('%')==-1:        # 순도가 95보다 낮은 경우에 문장에 포함된 순도를 찾아내 가져오는 부분
                for RealPurity in PurityArr:
                    if '%' in RealPurity:
                        Purity=RealPurity
                        break
                    else:
                        Purity=None
            if Purity is not None:
                PurityArr = Purity.split('%')
                Purity = float(int(PurityArr[0]) / 100)

        # Price
        GetID=soup.select_one('#commongrid > table > tbody > tr:nth-child(1) > td:nth-child(4)')        #물품의 고유 ID를 가져옴
        time.sleep(0.5)
        GetID=str(GetID).split('<')[2]
        GetID=(GetID.split('"')[1]).split('_price')[0]
        if driver.find_element_by_xpath('//*[@id="' + GetID + '_price' + '"]/div[1]').text!='Indicative prices':
            lenth=len(driver.find_element_by_xpath('//*[@id="' + GetID + '_price' + '"] / div / div[1] / div / div[1] / div / select').text.splitlines())
        else:
            lenth = len(driver.find_element_by_xpath('//*[@id="' + GetID + '_price' + '"]/div[2]/select').text.splitlines())     #Quantity / Price에 'Indicative prices'가 존재하는 물질들을 따로 처리하기 위한 부분(ex. 1205513-88-7)
        i = 1

        while i<lenth+1:
            istr=str(i)
            if driver.find_element_by_xpath('//*[@id="' + GetID + '_price' + '"]/div[1]').text != 'Indicative prices':
                xpath = '//*[@id="' + GetID + '_price' + '"]/div/div[1]/div/div[1]/div/select/option[' + istr + ']'
            else:
                xpath = '//*[@id="' + GetID + '_price' + '"]/div[2]/select/option[' + istr + ']'
            driver.find_element_by_xpath(xpath).click()
            Quantity=driver.find_element_by_xpath(xpath).text
            if driver.find_element_by_xpath('//*[@id="' + GetID + '_price' + '"]/div[1]').text != 'Indicative prices':
                if driver.find_element_by_xpath('//*[@id="' + GetID + '_price' + '"] / div / div[1] / p').text !='Out of Stock':
                    Price=driver.find_element_by_xpath('//*[@id="' + GetID + '_price' + '"]/div/div[1]/div/div[2]/p').text
                else:
                    Price=''
            else:
                Price = driver.find_element_by_xpath('//*[@id="' + GetID + '_price' + '"]/div[2]/div/span[1]').text
            WeightNum=Quantity.split(' ')[0]
            WeightUnit=Quantity.split(' ')[1]
            if math.floor(float(WeightNum)) == 0:
                Unit = str(int(float(WeightNum) * 1000)) + 'mg'
            elif math.floor(float(WeightNum))==float(WeightNum):
                Unit=str(int(float(WeightNum)))+WeightUnit
            else:
                Unit=str(WeightNum)+WeightUnit
            Price=Price.split('.0')[0]
            PriceAndUnitResult='USD'+Price+'/'+Unit
            PriceAndUnitList.append(PriceAndUnitResult)
            i+=1
        while j<=len(PriceAndUnitList)-1:
            if j==len(PriceAndUnitList)-1:
                PriceAndUnitStr=PriceAndUnitStr+PriceAndUnitList[j]
            else:
                PriceAndUnitStr=PriceAndUnitStr+PriceAndUnitList[j]+', '
            j+=1

        # Stock
        if driver.find_element_by_xpath(
                '//*[@id="commongrid"]/table/tbody/tr[1]/td[3]/span[1]').text=='Out of stock' and Purity is None:
            GlobalStock='Out of Stock'
        elif driver.find_element_by_xpath(
                '//*[@id="commongrid"]/table/tbody/tr[1]/td[3]/span[1]').text=='USA Rush delivery' and Purity is None:
            GlobalStock='USA Rush delivery Only'
        elif soup.select_one(ProductInfo + 'div:nth-child(2) > span.catdata1').text \
                == '\nMADE Building Block\n' and Purity is None:
            GlobalStock='Make-On-Demand Building Block'
        else:
            GlobalStock = driver.find_element_by_id('cartentry_null').text
            GlobalStock = GlobalStock.replace(' ', '')

        # 납기
        try:
            PeriodOfPayment = soup.select_one(Availability + 'span:nth-child(4)')
            if 'business days' in PeriodOfPayment.text:
                PeriodOfPayment = PeriodOfPayment.text
                PeriodOfPaymentArr = PeriodOfPayment.split('(')
                PeriodOfPayment = PeriodOfPaymentArr[1]
                PeriodOfPaymentArr = PeriodOfPayment.split(')')
                PeriodOfPayment = PeriodOfPaymentArr[0]
            elif GlobalStock == 'USA Rush delivery Only':
                PeriodOfPayment = soup.select_one(
                    '#commongrid > table > tbody > tr > td:nth-child(3) > span:nth-child(2) > em').text
                PeriodOfPaymentArr=PeriodOfPayment.split('(')
                PeriodOfPayment=PeriodOfPaymentArr[1]
                PeriodOfPaymentArr=PeriodOfPayment.split(')')
                PeriodOfPayment=PeriodOfPaymentArr[0]
            else:
                PeriodOfPayment = None

        except AttributeError:
            PeriodOfPayment = None
            if GlobalStock=='Make-On-Demand Building Block':
                PeriodOfPayment=soup.select_one(
                    '#commongrid > table > tbody > tr:nth-child(1) > td:nth-child(3) > span').text
            elif GlobalStock == 'Out of Stock':
                PeriodOfPayment = soup.select_one(
                    '#commongrid > table > tbody > tr:nth-child(1) > td:nth-child(3) > span.catdata.tooltip').text

        if driver.find_element_by_xpath('//*[@id="' + GetID + '_price' + '"] / div / div[1] / p').text == 'Out of Stock':
            PriceAndUnitStr=''
        Line=[Name,CASNum,CatalogID,Purity,PriceAndUnitStr,GlobalStock,PeriodOfPayment]
        excel_sheet1.append(Line)

        excel_file.save('EnamineCrawlingResult.xlsm')
        excel_file1=openpyxl.load_workbook('EnamineCrawlingResult.xlsm',read_only=False,keep_vba=True)
        excel_file1.save('EnamineCrawlingResult.xlsm')
        excel_file.close()
        excel_file1.close()         #엑셀파일 정리 및 저장 코드 집합

        EnaminePandasProcess.dropduplicates()

        driver.quit()
        break
    except PermissionError:
        print('엑셀 파일이 열려 있습니다. 파일을 닫고 다시 시도해 주세요.')

    except IndexError:
        data = driver.page_source
        soup = BeautifulSoup(data, 'html.parser')
        if soup.select_one('#content > div:nth-child(5) > p:nth-child(1)').text=='Nothing was found matching your criteria.':
            print('CAS 번호를 찾을 수 없습니다. 다시 확인해 주세요.')