import pandas as pd
import openpyxl
from openpyxl.styles import alignment

def dropduplicates():

    df=pd.read_excel('EnamineCrawlingResult.xlsm')
    df=df.drop_duplicates(subset='CAS Number',keep='last')      #데이터가 중복되면 이전에 저장된 데이터를 삭제하고 나중에 추가된 데이터를 남김

    df.to_excel('EnamineCrawlingResult.xlsm',index=False,header=False)

    excel_file=openpyxl.load_workbook('EnamineCrawlingResult.xlsm',read_only=False,keep_vba=True)
    excel_sheet1 = excel_file.active
    excel_sheet1.title='Catalog'
    excel_sheet1.column_dimensions['A'].width = 60
    excel_sheet1.column_dimensions['B'].width = 15
    excel_sheet1.column_dimensions['C'].width = 15
    excel_sheet1.column_dimensions['D'].width = 7
    excel_sheet1.column_dimensions['E'].width = 115
    excel_sheet1.column_dimensions['F'].width = 31
    excel_sheet1.column_dimensions['G'].width = 33
    excel_sheet1.insert_rows(1)
    excel_sheet1.cell(row=1,column=1).value='Name'
    excel_sheet1.cell(row=1,column=2).value='CAS Number'
    excel_sheet1.cell(row=1,column=3).value='Catalog Number'
    excel_sheet1.cell(row=1,column=4).value='Purity'
    excel_sheet1.cell(row=1,column=5).value='Price/Quantity'
    excel_sheet1.cell(row=1,column=6).value='Stock'
    excel_sheet1.cell(row=1,column=7).value='납기'
    i=1
    while i<=7:
        excel_sheet1.cell(row=1, column=i).alignment = openpyxl.styles.Alignment(horizontal='center')
        i+=1                    #엑셀 시트 변환 코드 집합
    excel_sheet1.protection.sheet=True
    excel_file.save('EnamineCrawlingResult.xlsm')
    excel_file.close()